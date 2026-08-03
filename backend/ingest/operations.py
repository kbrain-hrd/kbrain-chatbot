"""운영 자료를 마크다운으로 변환하고 섹션 앵커를 부여한다.

    uv run python -m backend.ingest.operations

문제 자료와 달리 형식이 제각각이라(hwpx · HTML · PDF) 추출기를 따로 둔다.
문항 구조가 없으므로 앵커는 `OPS-준비안내-응시환경` 처럼 문서-섹션 단위다.

멱등하다. 자료가 갱신되면 다시 돌리면 된다.
"""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path

from backend.ingest.extract import extract_pdfium

RAW = Path("raw")
OUT_DIR = Path("content/operations")

# hwpx 본문에서 `1. 개요` 처럼 번호가 붙은 문단이 섹션 경계다.
HWPX_HEADING_RE = re.compile(r"^(\d+)\.\s*(.+)$")


@dataclass
class Section:
    title: str
    body: str


@dataclass
class Document:
    slug: str  # 앵커에 쓰는 문서 이름 (예: 준비안내)
    source: Path
    sections: list[Section]


def slugify(text: str) -> str:
    """앵커에 쓸 수 있게 다듬는다. 한글은 그대로 두고 공백·기호만 정리."""
    cleaned = re.sub(r"[^\w가-힣]+", "", text)
    return cleaned or "본문"


# ---------------------------------------------------------------- hwpx


def extract_hwpx(path: Path) -> list[Section]:
    """hwpx 는 zip 안의 Contents/section0.xml 이 본문이다.

    텍스트가 서식 단위로 잘게 쪼개져 있어 문단(`<hp:p>`) 단위로 다시 붙인다.
    """
    with zipfile.ZipFile(path) as zf:
        xml = zf.read("Contents/section0.xml").decode("utf-8")

    paragraphs = []
    for chunk in re.findall(r"<hp:p\b.*?</hp:p>", xml, re.DOTALL):
        pieces = re.findall(r"<hp:t>(.*?)</hp:t>", chunk, re.DOTALL)
        text = "".join(re.sub(r"<[^>]+>", "", piece) for piece in pieces)
        text = re.sub(r"\s+", " ", text).strip()
        if text:
            paragraphs.append(text)

    return group_by_heading(paragraphs, HWPX_HEADING_RE)


def group_by_heading(paragraphs: list[str], heading_re: re.Pattern[str]) -> list[Section]:
    sections: list[Section] = []
    title = "머리말"
    buffer: list[str] = []

    for text in paragraphs:
        match = heading_re.match(text)
        if match:
            if buffer:
                sections.append(Section(title, "\n\n".join(buffer)))
                buffer = []
            title = match.group(2).strip()
        else:
            buffer.append(text)

    if buffer:
        sections.append(Section(title, "\n\n".join(buffer)))
    return sections


# ---------------------------------------------------------------- HTML


class GuideParser(HTMLParser):
    """h1/h2 를 섹션 경계로 삼아 본문 텍스트를 모은다.

    이 가이드는 페이지마다 같은 h1/h2 쌍을 반복 출력하므로, 제목이 바뀔 때만
    새 섹션을 연다. 같은 제목이 다시 나오면 이어 붙인다.
    """

    SKIP = {"script", "style", "head"}

    def __init__(self) -> None:
        super().__init__()
        self.sections: list[Section] = []
        self._heading_level: int | None = None
        self._heading_text: list[str] = []
        self._skip_depth = 0
        self._current: Section | None = None
        self._h1 = ""

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in self.SKIP:
            self._skip_depth += 1
        elif tag in ("h1", "h2"):
            self._heading_level = int(tag[1])
            self._heading_text = []

    def handle_endtag(self, tag: str) -> None:
        if tag in self.SKIP:
            self._skip_depth = max(0, self._skip_depth - 1)
        elif tag in ("h1", "h2") and self._heading_level is not None:
            title = re.sub(r"\s+", " ", "".join(self._heading_text)).strip()
            if self._heading_level == 1:
                self._h1 = title
            elif title:
                full = f"{self._h1} — {title}" if self._h1 else title
                if self._current is None or self._current.title != full:
                    self._current = Section(full, "")
                    self.sections.append(self._current)
            self._heading_level = None

    def handle_data(self, data: str) -> None:
        if self._skip_depth:
            return
        if self._heading_level is not None:
            self._heading_text.append(data)
            return
        text = re.sub(r"\s+", " ", data).strip()
        if text and self._current is not None:
            self._current.body += text + "\n"


def extract_html(path: Path) -> list[Section]:
    parser = GuideParser()
    parser.feed(path.read_text(encoding="utf-8", errors="replace"))
    # 반복 출력 탓에 같은 문장이 여러 번 담긴다. 순서를 지키며 중복 줄을 걷어낸다.
    for section in parser.sections:
        seen: set[str] = set()
        kept = []
        for line in section.body.splitlines():
            if line and line not in seen:
                seen.add(line)
                kept.append(line)
        section.body = "\n".join(kept)
    return [s for s in parser.sections if s.body.strip()]


# ---------------------------------------------------------------- PDF

# 로마숫자·장번호만 있거나 지나치게 짧은 줄은 섹션 제목이 될 수 없다
NOT_A_TITLE_RE = re.compile(r"[IVXⅠⅡⅢⅣⅤ\dI\s.·\-]{1,6}")


def running_headers(pages: list[list[str]]) -> set[str]:
    """페이지마다 반복되는 머리글을 찾는다.

    슬라이드형 PDF는 상단에 문서명·회차 같은 고정 문구를 매 장 반복한다. 그대로 두면
    그것이 섹션 제목으로 잡혀 실제 목차가 사라진다. 전체의 30% 이상 페이지에
    나타나는 앞부분 줄을 머리글로 본다.
    """
    from collections import Counter

    counts: Counter[str] = Counter()
    for lines in pages:
        for line in lines[:3]:
            counts[line] += 1

    threshold = max(2, len(pages) * 3 // 10)
    return {line for line, count in counts.items() if count >= threshold}


def extract_pdf(path: Path) -> list[Section]:
    """슬라이드형 PDF는 **페이지 상단 첫 줄이 제목** 역할을 한다.

    번호가 붙은 절 제목이 없으므로 페이지 제목으로 섹션을 나누고, 같은 제목이
    연달아 나오면(여러 장에 걸친 설명) 하나로 합친다.
    반복 머리글과 쪽번호만 있는 줄은 제목 후보에서 뺀다.
    """
    raw_pages = []
    for page in extract_pdfium(path):
        lines = [re.sub(r"\s+", " ", line).strip() for line in page.splitlines()]
        raw_pages.append([line for line in lines if line])

    headers = running_headers(raw_pages)
    sections: list[Section] = []

    for page_no, lines in enumerate(raw_pages, start=1):
        # 머리글·쪽번호를 걷어내고 남은 첫 줄이 그 장의 제목이다
        meaningful = [ln for ln in lines if ln not in headers and not re.fullmatch(r"[\d\-·.]{1,4}", ln)]
        if not meaningful:
            continue

        # 로마숫자·장번호만 있는 줄은 제목이 아니다. 뒤에 붙은 실제 제목을 찾는다.
        title_at = 0
        while title_at < len(meaningful) - 1 and NOT_A_TITLE_RE.fullmatch(meaningful[title_at]):
            title_at += 1

        title, rest = meaningful[title_at], meaningful[:title_at] + meaningful[title_at + 1 :]
        body = "\n".join([f"<!-- page {page_no} -->", *rest])

        if sections and sections[-1].title == title:
            sections[-1].body += "\n" + body
        else:
            sections.append(Section(title, body))

    return sections


# 운영 자료에는 담당자 이름·연락처가 섞여 있다. content/ 는 git 추적 대상이므로
# **출력 직전에 일괄 마스킹**한다. 추출기별로 처리하면 새 형식을 추가할 때 빠뜨린다.
# 직함 뒤를 함께 본다. `대리`·`사원` 은 일반명사이기도 해서 뒤를 보지 않으면
# "후배의 **대리 신청**", "불가피할 경우 **대리 가입**" 까지 이름으로 지운다.
# 실명은 `님`·조사·괄호를 달고 나오지만, 일반명사는 뒤에 명사가 이어진다.
NAME_WITH_TITLE_RE = re.compile(
    r"[가-힣]{2,4}\s*(?:주임|대리|과장|차장|부장|팀장|사원|매니저|강사|연구원|수석)"
    r"(?=님|께|이|가|은|는|을|를|에게|와|과|[)\]\},.]|$)"
)
PHONE_RE = re.compile(r"01[016-9][-.\s]?\d{3,4}[-.\s]?\d{4}")
TEL_RE = re.compile(r"0\d{1,2}[-.\s]\d{3,4}[-.\s]\d{4}")
EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.]+")


# 공개된 공식 문의처는 개인정보가 아니다. 지우면 "어디로 문의하나요" 에 답할 수 없다.
PUBLIC_CONTACTS = ("databus@nia.or.kr",)

# 자료 기준 연도. 해가 바뀌면 여기를 고친다 — `datetime.now()` 를 쓰면 실행 시점에 따라
# 산출물이 달라져 인제스트가 멱등하지 않게 된다.
CURRENT_YEAR = 2026

# "2024년 확정 정원" 처럼 **연도가 판정 기준으로 박힌** 서술. 그대로 답변에 실리면 오답이다.
# 단순히 연도가 등장하는 것(문제 지문 인용 등)까지 잡으면 15개 섹션이 걸려 표시가 무의미해진다.
STALE_YEAR_RE = re.compile(r"(20\d\d)년\s*(?:확정|기준|정원|이후|부터|까지|현재|말|상반기|하반기)")


def stale_years(body: str) -> list[str]:
    return sorted({y for y in STALE_YEAR_RE.findall(body) if y != str(CURRENT_YEAR)})


def scrub(text: str) -> str:
    text = EMAIL_RE.sub(lambda m: m.group() if m.group() in PUBLIC_CONTACTS else "[이메일]", text)
    text = PHONE_RE.sub("[연락처]", text)
    text = TEL_RE.sub("[연락처]", text)
    return NAME_WITH_TITLE_RE.sub("[담당자]", text)


def extract_txt(path: Path) -> list[Section]:
    """안내 메일 등 평문. `■` 로 시작하는 줄을 절 제목으로 본다."""
    text = path.read_text(encoding="utf-8", errors="replace")
    paragraphs = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    paragraphs = [p for p in paragraphs if p]
    return group_by_heading(paragraphs, re.compile(r"^■\s*()(.+)$"))


# 명단·연락처가 담긴 시트는 읽지 않는다
PII_SHEET_WORDS = ("명단", "연락처", "선발", "참석", "수강생", "교육생")


def extract_xlsx(path: Path) -> list[Section]:
    """운영 안내 문구가 담긴 시트만 읽는다. 개인 명단 시트는 건너뛴다."""
    from openpyxl import load_workbook

    sections: list[Section] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        if any(word in ws.title for word in PII_SHEET_WORDS):
            continue

        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [re.sub(r"\s+", " ", str(c)).strip() for c in row if c is not None]
            if cells:
                rows.append(" | ".join(cells))
        if rows:
            sections.append(Section(ws.title, "\n".join(rows)))
    wb.close()
    return sections


# 문의 이력 워크북에서 **읽어도 되는 시트**. 나머지 시트는 실명·소속·연락처가 담긴
# 상담 이력 8천 행이라 블랙리스트(PII_SHEET_WORDS)로는 시트가 늘 때마다 새어 나간다.
# 그래서 이 파일만은 화이트리스트로 막는다.
FAQ_SHEET = "AI챔피언자주하는 질문(25년~)"

# 응시자에게 나갈 답변이 아니라 담당자 업무 루틴이 적힌 행. 근거로 잡히면
# "진동으로 해놓고 부재중 확인" 같은 내부 지침을 인용하게 된다.
FAQ_INTERNAL_ROWS = {"문의 방법", "1:1 게시판/메일", "전화/문자"}


def extract_faq(path: Path) -> list[Section]:
    """문의 FAQ 시트 — **질문 한 줄이 섹션 하나**다.

    다른 운영 문서와 달리 주제로 묶지 않는다. 제목이 곧 질문 원문이라 그 자체가
    가장 정확한 조회 키이기 때문이다(`III AI 챔피언` 같은 무의미한 제목과 반대).
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if FAQ_SHEET not in wb.sheetnames:
        wb.close()
        return []

    sections = []
    for row in wb[FAQ_SHEET].iter_rows(values_only=True):
        cells = [re.sub(r"\s+", " ", str(c)).strip() for c in row[:2] if c is not None]
        if len(cells) < 2:
            continue  # 답변이 아직 없는 행
        question, answer = cells
        if question in FAQ_INTERNAL_ROWS:
            continue
        sections.append(Section(question, answer))

    wb.close()
    return sections


# ---------------------------------------------------------------- 출력


def render(doc: Document) -> str:
    out = [
        "---",
        f"doc: {doc.slug}",
        f"source: {doc.source.as_posix()}",
        f"sections: {len(doc.sections)}",
        "---",
        "",
        f"# {doc.slug}",
        "",
    ]
    # 간지처럼 같은 제목이 떨어져 다시 나오는 경우가 있다. 앵커가 겹치면 조회가
    # 모호해지므로 뒤에 순번을 붙여 유일하게 만든다.
    used: dict[str, int] = {}
    for section in doc.sections:
        # 마스킹을 먼저 한다. 제목만 가리고 앵커를 원문으로 만들면 앵커에 실명이 남는다.
        title = scrub(section.title)
        anchor = f"OPS-{doc.slug}-{slugify(title)}"
        used[anchor] = used.get(anchor, 0) + 1
        if used[anchor] > 1:
            anchor = f"{anchor}-{used[anchor]}"
        body = scrub(section.body).strip()
        # 답변으로 나가기 전에 사람이 확인해야 하는 대목을 본문에 남긴다. 초안 생성 때
        # 자료와 함께 읽히므로 모델도 이 단서를 보고 단정을 피할 수 있다.
        years = stale_years(body)
        if years:
            body += (
                f"\n\n<!-- 확인 필요: {'·'.join(years)}년 기준으로 서술된 내용이 있습니다. "
                f"현재({CURRENT_YEAR}년) 기준과 다를 수 있으니 그대로 답변하지 마세요. -->"
            )
        out += [f"## [{anchor}] {title}", "", body, ""]
    return "\n".join(out).rstrip() + "\n"


def find(pattern: str) -> Path | None:
    return next(iter(sorted(RAW.rglob(pattern))), None)


def build() -> list[Document]:
    candidates = [
        ("준비안내", RAW / "2026년_AI챔피언_자기주도형_수행평가_준비안내.hwpx"),
        ("셀프학습가이드", find("pkg/*/*.html")),
        ("CBT가이드", RAW / "CBT_가이드.pdf"),
        ("종합안내서", RAW / "교육과정_종합안내서_V1.3.pdf"),
        ("비대면교육사전안내", find("*비대면 교육 사전 안내*.pdf")),
        ("보조강사FAQ", find("*보조강사FAQ.xlsx")),
        ("운영준비체크리스트", find("*운영 준비 체크리스트.xlsx")),
        ("강사안내메일", find("*AI 리터러시와 업무활용 보조강사 안내 메일안.txt")),
        ("문의FAQ", find("*문의내용 정리*.xlsx")),
    ]

    # 확장자만으로 고를 수 없는 문서. 문의 이력 워크북은 xlsx 지만 통째로 읽으면
    # 개인정보가 쏟아지므로 FAQ 시트만 읽는 전용 추출기를 쓴다.
    overrides = {"문의FAQ": extract_faq}

    extractors = {
        ".hwpx": extract_hwpx,
        ".html": extract_html,
        ".pdf": extract_pdf,
        ".txt": extract_txt,
        ".xlsx": extract_xlsx,
    }

    docs = []
    for slug, source in candidates:
        if source is None or not source.is_file():
            print(f"  건너뜀 (파일 없음): {slug}")
            continue
        doc = Document(slug, source, [])
        extractor = overrides.get(slug) or extractors[source.suffix.lower()]
        doc.sections = extractor(source)
        docs.append(doc)

    return docs


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    docs = build()

    for doc in docs:
        path = OUT_DIR / f"{doc.slug}.md"
        path.write_text(render(doc), encoding="utf-8")
        size = path.stat().st_size
        print(f"{doc.slug:<12} 섹션 {len(doc.sections):>3}개  {size:>7,} bytes  → {path}")

    print(
        "\n종합안내서(60p)는 디자인 브로슈어라 읽기 순서가 뒤엉킵니다 "
        "— 추출 품질을 확인하고 쓸지 판단하세요 (docs/05-data-survey.md 5장)."
    )


if __name__ == "__main__":
    main()
