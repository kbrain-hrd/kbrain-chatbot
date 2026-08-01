"""수령한 질문 게시판 xlsx 에서 **질문·답변만** 추출한다.

    uv run python -m backend.sheets.extract

원본 시트에는 교육생 이름·소속기관·연락처·이메일이 들어 있다. 챗봇에 필요한 것은
질문 내용과 답변뿐이므로 개인정보 컬럼은 아예 읽지 않는다 (2026-08-01 결정).

컬럼을 버리는 것만으로는 부족하다 — 질문 본문에 연락처가 적히는 경우가 있어
전화번호·이메일 패턴을 마스킹하고, 그래도 남은 것이 있는지 스캔해 보고한다.

산출물은 3단계(질문 분류 + 문항 특정)의 골드셋 기반이 된다.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

SHEETS_ROOT = Path("raw/sheets")
OUT_PATH = Path("eval/questions.md")

QUESTION_HEADERS = {"질문 내용", "질문"}
# 읽지 않을 컬럼 — 개인정보
PII_HEADERS = {"이름", "교육생 이름", "소속기관", "소속기관구분", "연락처", "이메일", "전화번호"}

PHONE_RE = re.compile(r"01[016-9][-.\s]?\d{3,4}[-.\s]?\d{4}")
EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.]+")
# 기관 대표번호 등도 개인 식별로 이어질 수 있어 함께 본다
TEL_RE = re.compile(r"0\d{1,2}[-.\s]\d{3,4}[-.\s]\d{4}")


@dataclass
class Entry:
    source: str
    date: str
    category: str
    question: str
    status: str
    answer: str


def mask(text: str) -> str:
    text = EMAIL_RE.sub("[이메일]", text)
    text = PHONE_RE.sub("[연락처]", text)
    return TEL_RE.sub("[연락처]", text)


def clean(value: object) -> str:
    if value is None:
        return ""
    return mask(re.sub(r"\s+", " ", str(value)).strip())


def read_entries(path: Path) -> list[Entry]:
    entries: list[Entry] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"  열기 실패 {path.name}: {exc}")
        return entries

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))

        header_at = None
        for index, row in enumerate(rows[:12]):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if QUESTION_HEADERS & set(cells):
                header_at, header = index, cells
                break
        if header_at is None:
            continue

        def at(*names: str) -> int | None:
            for name in names:
                if name in header:
                    return header.index(name)
            return None

        q_at = at(*QUESTION_HEADERS)
        date_at = at("작성일", "시각")
        cat_at = at("회차", "과정명")
        status_at = at("상태")
        answer_at = at("답변")

        def cell(row: tuple, index: int | None) -> str:
            if index is None or index >= len(row):
                return ""
            return clean(row[index])

        for row in rows[header_at + 1 :]:
            question = cell(row, q_at)
            if not question:
                continue
            entries.append(
                Entry(
                    source=f"{path.stem} [{ws.title}]",
                    date=cell(row, date_at),
                    category=cell(row, cat_at),
                    question=question,
                    status=cell(row, status_at),
                    answer=cell(row, answer_at),
                )
            )
    wb.close()
    return entries


def scan_residual(entries: list[Entry]) -> list[tuple[str, str]]:
    """마스킹 후에도 남은 개인정보 신호를 찾는다."""
    hits = []
    for entry in entries:
        for field, text in (("질문", entry.question), ("답변", entry.answer)):
            if PHONE_RE.search(text) or EMAIL_RE.search(text) or TEL_RE.search(text):
                hits.append((entry.source, f"{field}: {text[:60]}"))
    return hits


def render(entries: list[Entry]) -> str:
    out = [
        "# 실제 질문 모음 (골드셋 기반)",
        "",
        "`uv run python -m backend.sheets.extract` 산출물.",
        "원본 시트의 **개인정보 컬럼(이름·소속기관·연락처·이메일)은 읽지 않는다.**",
        "본문에 남은 전화번호·이메일 패턴은 `[연락처]` · `[이메일]` 로 마스킹한다.",
        "",
        f"총 {len(entries)}건. 3단계(질문 분류 + 문항 특정)의 정확도 측정에 쓴다.",
        "",
    ]

    by_source: dict[str, list[Entry]] = {}
    for entry in entries:
        by_source.setdefault(entry.source, []).append(entry)

    for source, items in by_source.items():
        out += [f"## {source}", "", f"{len(items)}건", ""]
        for index, entry in enumerate(items, start=1):
            head = " · ".join(filter(None, [entry.date, entry.category, entry.status]))
            out += [f"### {index}. {head}" if head else f"### {index}."]
            out += ["", f"**질문:** {entry.question}", ""]
            if entry.answer:
                out += [f"**답변:** {entry.answer}", ""]
    return "\n".join(out).rstrip() + "\n"


def main() -> None:
    entries: list[Entry] = []
    for path in sorted(SHEETS_ROOT.rglob("*.xlsx")):
        entries.extend(read_entries(path))

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(render(entries), encoding="utf-8")

    size = OUT_PATH.stat().st_size
    print(f"질문 {len(entries)}건 → {OUT_PATH} ({size:,} bytes)")

    residual = scan_residual(entries)
    if residual:
        print(f"\n마스킹 후에도 개인정보 신호가 남은 항목 {len(residual)}건 — 확인 필요:")
        for source, text in residual[:20]:
            print(f"  - {source} {text}")
    else:
        print("본문에서 전화번호·이메일 패턴 없음.")


if __name__ == "__main__":
    main()
