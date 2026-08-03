"""문의 이력 xlsx 에서 질문·답변 쌍만 뽑는다.

    uv run python -m backend.sheets.inquiries

`raw/[문의내용 정리] ….xlsx` 는 13개 시트 8,694행이고 그중 8,383행이 **실명·소속·
휴대폰번호가 담긴 상담 이력**이다. 그래서 두 겹으로 막는다.

1. **읽을 컬럼을 이름으로 지정한다.** 문의내용·답변·처리현황만 읽고 이름·소속·연락처
   컬럼은 아예 건드리지 않는다.
2. **본문도 마스킹한다.** 응시자가 본문에 "저는 ○○기관 김□□인데 010-…" 이라고 적는
   경우가 있어 컬럼을 걸러도 남는다.

산출물은 git 제외다. 마스킹 후에도 소속기관이 문맥으로 드러날 수 있다.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from backend.ingest.operations import scrub

RAW = Path("raw")
OUT_PATH = Path("eval/inquiries.md")

# 읽을 컬럼. 부분 일치로 찾는다 — 시트마다 `연락처`/`전화번호` 처럼 표기가 다르다.
QUESTION_COLUMNS = ("문의내용",)
ANSWER_COLUMNS = ("답변",)
STATUS_COLUMNS = ("처리현황",)

# 상담 이력이 담긴 시트. FAQ 시트(`AI챔피언자주하는 질문`)는 이미 인제스트에 들어가 있다.
HISTORY_SHEETS = (
    "2-1.문의사항(Call)",
    "2-2. 문의사항(Mail)",
    "2-3. 문의사항(11)",
    "252-1. 문의사항(Call)",
    "252-2. 문의사항(Maill)",
    "252-3. 문의사항(11)",
)

# AI 챔피언·셀프스터디와 무관한 일반 교육 문의가 대부분이라 걸러 낸다.
TOPIC_RE = re.compile(r"챔피언|인증평가|셀프|CBT|그린|블루|웹캠|수행평가|자기주도|기관맞춤")

# 답변이 아직 없는 행은 FAQ 가 될 수 없다
UNANSWERED = ("확인중", "처리중", "NIA 요청", "미처리")


@dataclass(frozen=True)
class Inquiry:
    sheet: str
    question: str
    answer: str


def column_map(header: tuple, *names: str) -> int | None:
    for index, cell in enumerate(header):
        if cell and any(name in str(cell) for name in names):
            return index
    return None


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value)).strip() if value is not None else ""


def read_inquiries(path: Path) -> list[Inquiry]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    found: list[Inquiry] = []

    for name in HISTORY_SHEETS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        q_at = a_at = s_at = None

        for row in ws.iter_rows(values_only=True):
            if q_at is None:
                q_at = column_map(row, *QUESTION_COLUMNS)
                a_at = column_map(row, *ANSWER_COLUMNS)
                s_at = column_map(row, *STATUS_COLUMNS)
                continue

            question = clean(row[q_at]) if q_at < len(row) else ""
            answer = clean(row[a_at]) if a_at is not None and a_at < len(row) else ""
            status = clean(row[s_at]) if s_at is not None and s_at < len(row) else ""

            if len(question) < 15 or len(answer) < 10:
                continue
            if status in UNANSWERED:
                continue
            if not TOPIC_RE.search(f"{question} {answer}"):
                continue

            found.append(Inquiry(name, scrub(question), scrub(answer)))

    wb.close()
    return found


def dedupe(items: list[Inquiry]) -> list[Inquiry]:
    seen: set[tuple[str, str]] = set()
    kept = []
    for item in items:
        key = (item.question[:120], item.answer[:120])
        if key in seen:
            continue
        seen.add(key)
        kept.append(item)
    return kept


def find_source() -> Path | None:
    return next(iter(sorted(RAW.glob("*문의내용 정리*.xlsx"))), None)


def render(items: list[Inquiry]) -> str:
    out = [
        "# 문의 이력 (질문·답변만)",
        "",
        "`backend/sheets/inquiries.py` 산출물. **git 제외** — 마스킹 후에도 소속기관이",
        "문맥으로 드러날 수 있다. 이름·소속·연락처 컬럼은 애초에 읽지 않는다.",
        "",
        f"총 {len(items)}건.",
        "",
    ]
    for index, item in enumerate(items, start=1):
        out += [
            f"## {index}. [{item.sheet}]",
            "",
            f"**문의** {item.question}",
            "",
            f"**답변** {item.answer}",
            "",
        ]
    return "\n".join(out)


def main() -> None:
    source = find_source()
    if source is None:
        raise SystemExit("raw/ 에 문의내용 정리 xlsx 가 없습니다.")

    items = dedupe(read_inquiries(source))
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(render(items), encoding="utf-8")

    by_sheet: dict[str, int] = {}
    for item in items:
        by_sheet[item.sheet] = by_sheet.get(item.sheet, 0) + 1

    print(f"질문·답변 쌍 {len(items)}건 → {OUT_PATH}")
    for sheet, count in sorted(by_sheet.items(), key=lambda kv: -kv[1]):
        print(f"  {sheet:<24} {count:>4}")


if __name__ == "__main__":
    main()
